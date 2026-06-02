"""Excel import, column mapping, version diff API routes."""
import os
import uuid
import json
from datetime import datetime
from typing import List

from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.core.database import get_db
from app.core.config import settings
from app.models.ship_config import (
    ModbusConfig, SignalPoint, ConfigVersion, VersionChange, ImportTemplate,
)
from app.schemas.imports import (
    ImportTemplateCreate, ImportTemplateResponse,
    VersionDiffResponse, BatchApproveRequest,
)

router = APIRouter(prefix="/api/import", tags=["import"])


@router.post("/preview")
async def preview_excel(
    file: UploadFile = File(...),
    header_row: int = Form(7),
):
    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
    ext = os.path.splitext(file.filename)[1]
    save_path = os.path.join(settings.UPLOAD_DIR, f"{uuid.uuid4()}{ext}")
    content = await file.read()
    with open(save_path, "wb") as f:
        f.write(content)

    wb = load_workbook(save_path, data_only=True)
    ws = wb.active

    headers = []
    if header_row <= ws.max_row:
        headers = [str(c.value or "") for c in ws[header_row]]

    rows = []
    max_col = len(headers) or ws.max_column
    for row_idx in range(header_row + 1, min(header_row + 21, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            row_data.append(str(val) if val is not None else "")
        rows.append(row_data)

    return {
        "file_path": save_path,
        "file_name": file.filename,
        "sheet_name": ws.title,
        "total_rows": ws.max_row,
        "total_cols": max_col,
        "headers": headers,
        "rows": rows,
    }


@router.post("/templates", response_model=ImportTemplateResponse)
def create_template(req: ImportTemplateCreate, db: Session = Depends(get_db)):
    template = ImportTemplate(
        name=req.name, manufacturer=req.manufacturer,
        ship_series=req.ship_series, description=req.description,
        column_mapping=req.column_mapping.dict(),
        header_row=req.header_row, data_start_row=req.data_start_row,
        comm_params_row=req.comm_params_row,
    )
    db.add(template); db.commit(); db.refresh(template)
    return template


@router.get("/templates", response_model=List[ImportTemplateResponse])
def list_templates(db: Session = Depends(get_db)):
    return db.query(ImportTemplate).all()


@router.get("/templates/{template_id}", response_model=ImportTemplateResponse)
def get_template(template_id: int, db: Session = Depends(get_db)):
    template = db.query(ImportTemplate).filter(ImportTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    return template


@router.put("/templates/{template_id}")
def update_template(template_id: int, req: ImportTemplateCreate,
                    db: Session = Depends(get_db)):
    template = db.query(ImportTemplate).filter(ImportTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    template.name = req.name; template.manufacturer = req.manufacturer
    template.ship_series = req.ship_series; template.description = req.description
    template.column_mapping = req.column_mapping.dict()
    template.header_row = req.header_row; template.data_start_row = req.data_start_row
    template.comm_params_row = req.comm_params_row
    db.commit()
    return {"message": "已更新"}


@router.delete("/templates/{template_id}")
def delete_template(template_id: int, db: Session = Depends(get_db)):
    template = db.query(ImportTemplate).filter(ImportTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    db.delete(template); db.commit()
    return {"message": "已删除"}


@router.post("/parse")
def parse_excel(file_path: str = Form(...), template_id: int = Form(...),
                modbus_config_id: int = Form(...), db: Session = Depends(get_db)):
    template = db.query(ImportTemplate).filter(ImportTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    config = db.query(ModbusConfig).filter(ModbusConfig.id == modbus_config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail="Modbus配置不存在")

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    mapping = template.column_mapping

    def col_to_idx(letter):
        idx = 0
        for c in letter:
            idx = idx * 26 + (ord(c.upper()) - ord("A") + 1)
        return idx

    points = []
    for row_idx in range(template.data_start_row, ws.max_row + 1):
        name_val = ws.cell(row=row_idx, column=col_to_idx(mapping.get("name", "B"))).value
        addr_val = ws.cell(row=row_idx, column=col_to_idx(mapping.get("modbus_address", "C"))).value
        if not name_val or not addr_val:
            continue
        name = str(name_val).strip()
        if not name:
            continue

        point = SignalPoint(
            modbus_config_id=modbus_config_id,
            name=name,
            modbus_address=str(addr_val).strip(),
            signal_type=str(ws.cell(row=row_idx, column=col_to_idx(mapping.get("signal_type", "D"))).value or "").strip() or "开关量",
        )

        for field, col_letter in mapping.items():
            if field in ("name", "modbus_address", "signal_type", "point_index"):
                continue
            if col_letter:
                val = ws.cell(row=row_idx, column=col_to_idx(col_letter)).value
                if val is not None:
                    setattr(point, field, str(val).strip())

        idx_val = ws.cell(row=row_idx, column=col_to_idx(mapping.get("point_index", "A"))).value
        if idx_val is not None:
            try:
                point.point_index = int(float(str(idx_val)))
            except ValueError:
                pass

        points.append(point)

    for p in points:
        db.add(p)
    db.commit()
    return {"success": True, "message": f"已导入 {len(points)} 个工况点", "points_count": len(points)}


@router.post("/diff")
def create_version_and_diff(modbus_config_id: int = Form(...),
                            version_tag: str = Form(...), source_file: str = Form(""),
                            db: Session = Depends(get_db)):
    latest_version = db.query(ConfigVersion).filter(
        ConfigVersion.modbus_config_id == modbus_config_id
    ).order_by(ConfigVersion.id.desc()).first()

    current_points = db.query(SignalPoint).filter(
        SignalPoint.modbus_config_id == modbus_config_id
    ).all()

    new_version = ConfigVersion(
        modbus_config_id=modbus_config_id, version_tag=version_tag,
        source_file=source_file, created_by=None,
    )
    db.add(new_version); db.flush()

    for p in current_points:
        p.version_id = new_version.id

    added, deleted, modified = [], [], []

    if latest_version:
        old_points = db.query(SignalPoint).filter(
            SignalPoint.version_id == latest_version.id
        ).all()

        old_map = {p.ams_module_id or p.modbus_address: p for p in old_points}
        new_map = {p.ams_module_id or p.modbus_address: p for p in current_points}

        for key in set(new_map.keys()) - set(old_map.keys()):
            p = new_map[key]
            added.append({"name": p.name, "modbus_address": p.modbus_address,
                          "signal_type": p.signal_type, "ams_module_id": p.ams_module_id})

        for key in set(old_map.keys()) - set(new_map.keys()):
            p = old_map[key]
            deleted.append({"name": p.name, "modbus_address": p.modbus_address,
                            "signal_type": p.signal_type, "ams_module_id": p.ams_module_id})

        for key in set(old_map.keys()) & set(new_map.keys()):
            old_p = old_map[key]; new_p = new_map[key]
            diff_fields = []
            for attr in ("name", "modbus_address", "signal_type", "unit", "data_type"):
                old_val = getattr(old_p, attr) or ""
                new_val = getattr(new_p, attr) or ""
                if old_val != new_val:
                    diff_fields.append({"field": attr, "old_value": old_val, "new_value": new_val})
                    db.add(VersionChange(version_id=new_version.id, point_id=new_p.id,
                                         change_type="modified", field_name=attr,
                                         old_value=str(old_val), new_value=str(new_val)))
            if diff_fields:
                modified.append({"name": new_p.name, "modbus_address": new_p.modbus_address, "diff_fields": diff_fields})

    for item in added:
        db.add(VersionChange(version_id=new_version.id, change_type="added",
                             new_value=json.dumps(item, ensure_ascii=False)))
    for item in deleted:
        db.add(VersionChange(version_id=new_version.id, change_type="deleted",
                             old_value=json.dumps(item, ensure_ascii=False)))

    db.commit()
    return VersionDiffResponse(
        old_version=latest_version.version_tag if latest_version else "N/A",
        new_version=version_tag, added=added, deleted=deleted, modified=modified,
    )


@router.get("/changes/{version_id}")
def list_changes(version_id: int, db: Session = Depends(get_db)):
    changes = db.query(VersionChange).filter(VersionChange.version_id == version_id).all()
    return [{"id": c.id, "change_type": c.change_type, "field_name": c.field_name,
             "old_value": c.old_value, "new_value": c.new_value, "status": c.status} for c in changes]


@router.post("/changes/batch-approve")
def batch_approve(req: BatchApproveRequest, db: Session = Depends(get_db)):
    query = db.query(VersionChange)
    if req.change_ids:
        query = query.filter(VersionChange.id.in_(req.change_ids))
    new_status = "accepted" if req.action == "accept_all" else "rejected"
    now = datetime.now()
    for change in query.all():
        change.status = new_status
        change.approved_at = now
    db.commit()
    return {"message": f"已{new_status}变更"}


@router.post("/cfg2db/{config_id}")
def cfg2db(config_id: int,
           target_host: str = Form("192.168.8.48"), target_port: int = Form(3306),
           target_user: str = Form("root"), target_password: str = Form(""),
           target_db: str = Form("hlx"), db: Session = Depends(get_db)):
    try:
        import pymysql
        conn = pymysql.connect(host=target_host, port=target_port,
                               user=target_user, password=target_password,
                               database=target_db, connect_timeout=5)
        cursor = conn.cursor()

        config = db.query(ModbusConfig).filter(ModbusConfig.id == config_id).first()
        if not config:
            raise HTTPException(status_code=404, detail="配置不存在")
        points = db.query(SignalPoint).filter(
            SignalPoint.modbus_config_id == config_id, SignalPoint.is_active == 1
        ).all()

        results = []
        for p in points:
            results.append({"Addr": p.modbus_address, "Alias": p.ams_module_id or p.modbus_address,
                            "Coefficient": p.coefficient or "0,1,0", "DataLen": "1", "DataTypeRegister": "2"})

        json_string = json.dumps({
            "ChanCode": str(config.channel_id), "ChanDesc": config.name,
            "ChanType": "7", "DataProtocol": "6",
            "NetModTcp": [{"DataLen": str(len(points)), "DevName": config.name,
                           "FunCode": str(config.func_code_read_ai),
                           "RegisterBegin": str(points[0].modbus_address) if points else "40001",
                           "Results": results, "SlaveId": str(config.slave_id).zfill(2)}],
            "WayUid": "",
        }, ensure_ascii=False)

        cursor.execute("""INSERT INTO channelsdataprotocol
            (ShipId, WayUid, ChanType, ChanCode, DataProtocol, ChanDesc, JsonString)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE JsonString=VALUES(JsonString)""",
            ("1", "", 7, config.channel_id, 6, config.name, json_string))
        conn.commit(); cursor.close(); conn.close()
        return {"success": True, "message": f"已入库 {len(points)} 个工况点"}
    except Exception as e:
        return {"success": False, "message": f"入库失败: {str(e)}"}
